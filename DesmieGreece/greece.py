from datetime import datetime
import os
import shutil
import time
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.common.by import By

# Automatic downloads
download_dir = "C:\\Users\\User\\Desktop\\Python\\MVM\\PEX\\DesmieGreece\\Files" # Download directory

chrome_options = Options()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": download_dir,  # Default dir
    "download.prompt_for_download": False,       # No prompt for download
    "download.directory_upgrade": True,          # Auto overwriting
    "safebrowsing.enabled": True                 # Safe browsing
})

# Set up WebDriver
service = Service('C:\\Users\\User\\Desktop\\chromedriver-win64\\chromedriver.exe')
driver = webdriver.Chrome(service=service, options=chrome_options)

wait = WebDriverWait(driver, 20)

# Open the webpage
driver.get("https://www.enexgroup.gr/web/guest/markets-publications-el-day-ahead-market")  # Url of the website


# Cookies popup
cookie_button = driver.find_element(By.CLASS_NAME, 'btn-primary')
cookie_button.click()


downloads = driver.find_elements(By.XPATH, f'//*[@class="portlet-body"]//a[contains(., "ResultsSummary")]') # Elements do be clicked/downloaded
#for d in downloads:
#    print(d.text)

# First always delete all files from the download_dir folder
for filename in os.listdir(download_dir):
    file_path = os.path.join(download_dir, filename)
    try:
        if os.path.isfile(file_path):  # Check if it's a file
            os.remove(file_path)  # Delete the file
            print(f"Deleted: {file_path}")
    except Exception as e:
        print(f"Failed to delete {file_path}: {e}")

# Click on the link - then wait a bit
for i, button in enumerate(downloads):
    button.click()
    time.sleep(5)


# Rename downloaded files
new_file_names = [f"file_{i+1}.xlsx" for i in range(len(downloads))]
for i, filename in enumerate(os.listdir(download_dir)):
    old_path = os.path.join(download_dir, filename)
    new_path = os.path.join(download_dir, new_file_names[i])
    if os.path.isfile(old_path):
        shutil.move(old_path, new_path)
        print(f"Renamed: {old_path} to {new_path}")


# Needed data
buy = []
sell = []
buy_sell_max = []
market_clearing_price = []
time_column = []

# Get a list of all Excel files in the folder
excel_files = [file for file in os.listdir(download_dir) if file.endswith(('.xlsx', '.xlsm'))]

# Process files in chunks of 7
batch_size = 7
for batch_start in range(0, len(excel_files), batch_size):
    batch_files = excel_files[batch_start:batch_start + batch_size]
    print(f"Processing batch: {batch_files}")
    
    for file_name in batch_files:
        file_path = os.path.join(download_dir, file_name)
        print(f"Processing file: {file_name}")
        
        # Open the Excel file
        workbook = load_workbook(file_path)
        time_data = []
        i = 0
        for sheet_name in workbook.sheetnames[:2]:
            sheet = workbook[sheet_name]
            
            time_data = [str(sheet.cell(row=2, column=col).value) for col in range(1, 26)]
            
            if i == 0:
                # Extract data from row 4, columns B through Y
                buy.extend([sheet.cell(row=4, column=col).value for col in range(2, 26)])  # Columns B (2) to Y (25)
            else:
                sell.extend([sheet.cell(row=4, column=col).value for col in range(2, 26)])
                market_clearing_price.extend([sheet.cell(row=7, column=col).value for col in range(2, 26)])
            
            i += 1
        
        parsed_date = datetime.strptime(time_data[0], "%Y-%m-%d %H:%M:%S")  # Parse the date string
        day_month = parsed_date.strftime("%d %b")  # Format as '16 Dec'

        for i in range(1, len(time_data)):
            if int(time_data[i]) < 10:
                if time_data[i] == '1':
                    time_column.append(f'{day_month} 00 - 0{time_data[i]}')
                else:
                    time_column.append(f'{day_month} 0{time_data[i-1]} - 0{time_data[i]}')
            elif int(time_data[i]) == 10:
                time_column.append(f'{day_month} 0{time_data[i-1]} - {time_data[i]}')
            else:
                time_column.append(f'{day_month} {time_data[i-1]} - {time_data[i]}')

for i in range(len(buy)):
    buy_sell_max.append(max(buy[i], sell[i]))

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Add a header to the worksheet
ws.append(["Time", "Market Clearing Price", "Max of BUY and SELL"])  # Adds "Time" and "Value" as the header
#print(len(time_column), "AAA", len(market_clearing_price), "AAA", len(buy_sell_max))
# Write data into the worksheet
for time_value, mcp, max in zip(time_column, market_clearing_price, buy_sell_max):
    #print(time_value, mcp, max)
    ws.append([time_value, mcp, max])  # Each value goes into a new row

# Save the workbook
wb.save("stored_data.xlsx")